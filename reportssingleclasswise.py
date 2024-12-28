import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Load the Excel file without assuming any headers
file_path = 'class1.xlsx'  # Replace with your actual file path
df = pd.read_excel(file_path, header=None)

# List of required columns
required_columns = ['S.No', 'Student Name', 'Roll No', 'Enrollment Code', 'Class - Section', 'Admission No',
                    'English', 'Telugu', 'Mathematics', 'Science','Computer']

# Identify the header row
header_row_index = None
for i, row in df.iterrows():
    if set(required_columns).issubset(set(row)):
        header_row_index = i
        break

if header_row_index is not None:
    print(f"Header row found at index {header_row_index}. Reading data...")
    
    # Read the Excel file again using the identified header row
    df = pd.read_excel(file_path, header=header_row_index)

    # Convert subject columns to numeric, coerce errors to NaN
    subject_columns = required_columns[6:]
    df[subject_columns] = df[subject_columns].apply(pd.to_numeric, errors='coerce')

    # Replace NaN values in subject columns with 'ab'
    df[subject_columns] = df[subject_columns].fillna('ab')

    # Convert 'Admission No' to string to ensure consistent data type
    df['Admission No'] = df['Admission No'].astype(str)

    # Calculate total marks
    df['Total Marks'] = df[subject_columns].apply(pd.to_numeric, errors='coerce').sum(axis=1)

    # Rank students within each section
    df['Rank'] = df.groupby('Class - Section')['Total Marks'].rank(ascending=False, method='min')

    # Sort by 'Admission No'
    df = df.sort_values(by='Admission No')

    # Save the output files section-wise
    sections = df['Class - Section'].unique()
    for section in sections:
        section_df = df[df['Class - Section'] == section].copy()
        
        # Add 'S.No' column starting from 1
        section_df['S.No'] = range(1, len(section_df) + 1)

        # Remove 'Roll No' and 'Enrollment Code' columns if they exist
        columns_to_remove = ['Roll No', 'Enrollment Code']
        existing_columns = [col for col in columns_to_remove if col in section_df.columns]
        section_df = section_df[[col for col in section_df.columns if col not in existing_columns]]

        # Ensure only required columns are kept
        final_columns = [col for col in required_columns[:6] + subject_columns + ['Total Marks', 'Rank', 'S.No'] if col in section_df.columns]
        section_df = section_df[final_columns]

        # Remove any completely empty rows
        section_df = section_df.dropna(how='all')

        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Add header
        header1 = "VRS & VJ Residential School"
        header2 = f"Exam Marks Report for {section} Periodic Test-1"
        
        ws.append([header1])
        ws.append([header2])
        
        # Merge header cells
        ws.merge_cells('A1:L1')  # Adjust range if needed
        ws.merge_cells('A2:L2')  # Adjust range if needed

        # Apply header formatting
        header_font = Font(bold=True)
        ws['A1'].font = header_font
        ws['A2'].font = header_font

        # Append the DataFrame to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(section_df, index=False, header=True), 3):
            ws.append(row)

        # Set alignment for header
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set alignment for data cells
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.column < len(subject_columns) + 6:  # Numeric columns based on column index
                    cell.alignment = Alignment(horizontal='center')
                else:  # Text columns
                    cell.alignment = Alignment(horizontal='left')

        # Set borders for the entire data range
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = thin_border

        # Add footer: Class Teacher (left), Principal (right)
        ws.oddFooter.left.text = "Class Teacher"
        ws.oddFooter.right.text = "Principal"

        # Set page layout to A4, landscape mode, narrow margins
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # Autofit column width
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Remove any extra columns
        last_col = len(section_df.columns)
        ws.delete_cols(last_col)

        # Save the Excel sheet
        excel_file = f'section_{section}_ranked.xlsx'
        wb.save(excel_file)

    print("Files saved successfully.")
else:
    print("The required columns were not found in the file.")
