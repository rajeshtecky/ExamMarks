import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Define input and output directories
input_dir = r'D:\rajesh\examreports\test'
bg_dir = r'D:\rajesh\examreports\bg'
output_dir = r'D:\rajesh\examreports\output'

# Ensure the output directory exists
os.makedirs(output_dir, exist_ok=True)

# Load the BG marks file
bg_file = os.path.join(bg_dir, 'class5.xlsx')
bg_df = pd.read_excel(bg_file, header=None)

# List of required columns for BG marks
bg_columns = ['Admission No', 'BG']

# Identify the header row in BG file
header_row_index = None
for i in range(len(bg_df)):  # Check all rows for the header
    if set(bg_columns).issubset(set(bg_df.iloc[i].dropna())):
        header_row_index = i
        break

if header_row_index is not None:
    print(f"BG Header row found at index {header_row_index}. Reading data...")
    bg_df = pd.read_excel(bg_file, header=header_row_index)
    bg_df.columns = bg_df.columns.str.strip()  # Remove any leading/trailing spaces
    print("BG DataFrame columns:", bg_df.columns)  # Print column names for debugging
    bg_df = bg_df[bg_columns]
    bg_df['Admission No'] = bg_df['Admission No'].astype(str)  # Ensure Admission No is string
else:
    print("Header row for BG marks not found. Available rows:")
    print(bg_df.head())  # Print the first few rows for debugging
    raise ValueError("Header row for BG marks not found.")

# Process each file in the input directory
for filename in os.listdir(input_dir):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(input_dir, filename)
        print(f"Processing file: {file_path}")

        # Load the Excel file without assuming any headers
        df = pd.read_excel(file_path, header=None)
        print(f"Preview of file '{filename}':")
        print(df.head())  # Print the first few rows for debugging

        # Identify the header row
        header_row_index = None
        for i in range(len(df)):  # Check all rows for the header
            if set(df.iloc[i].dropna()).issubset(set(df.columns)):
                header_row_index = i
                break

        if header_row_index is not None:
            print(f"Header row found at index {header_row_index}. Reading data...")
            df = pd.read_excel(file_path, header=header_row_index)
            df.columns = df.columns.str.strip()  # Remove any leading/trailing spaces
            print("Columns in DataFrame:", df.columns)  # Print column names for debugging

            if 'Admission No' not in df.columns:
                print(f"'Admission No' column not found in {filename}. Available columns:")
                print(df.columns)  # Print column names for debugging
                continue  # Skip this file and move to the next

            df['Admission No'] = df['Admission No'].astype(str)  # Ensure Admission No is string

            # Identify available subjects
            subject_columns = [col for col in df.columns if col not in ['S.No', 'Student Name', 'Roll No', 'Enrollment Code', 'Class - Section', 'Admission No']]

            if 'ArtificialIntelligence' in subject_columns or 'Computer' in subject_columns:
                if 'ArtificialIntelligence' in subject_columns:
                    subject_columns.remove('ArtificialIntelligence')
                if 'Computer' in subject_columns:
                    subject_columns.remove('Computer')

            # Ensure BG Marks column is included
            if 'BG' not in subject_columns:
                subject_columns.append('BG')

            df = df[['Admission No'] + subject_columns]
            df = df.merge(bg_df, on='Admission No', how='left')

            # Add 'BG Marks' to columns and update total marks calculation
            subject_columns = [col for col in subject_columns if col != 'Admission No'] + ['BG']
            df[subject_columns] = df[subject_columns].apply(pd.to_numeric, errors='coerce')

            # Replace NaN values in subject columns with 'ab'
            df[subject_columns] = df[subject_columns].fillna('ab')

            # Calculate total marks including BG Marks
            df['Total Marks'] = df[subject_columns].apply(pd.to_numeric, errors='coerce').sum(axis=1)

            # Rank students within each section
            df['Rank'] = df.groupby('Class - Section')['Total Marks'].rank(ascending=False, method='min')

            # Sort by 'Admission No'
            df = df.sort_values(by='Admission No')

            # Create a folder for the current file
            file_output_dir = os.path.join(output_dir, os.path.splitext(filename)[0])
            os.makedirs(file_output_dir, exist_ok=True)

            # Check if DataFrame is not empty
            if df.empty:
                print(f"No data to save for {filename}")
                continue

            # Save the output files section-wise
            sections = df['Class - Section'].unique()
            for section in sections:
                section_df = df[df['Class - Section'] == section].copy()

                # Add 'S.No' column starting from 1
                section_df['S.No'] = range(1, len(section_df) + 1)

                # Remove 'Roll No' and 'Enrollment Code' columns if they exist
                columns_to_remove = ['Roll No', 'Enrollment Code']
                existing_columns = [col for col in columns_to_remove if col in section_df.columns]
                section_df = section_df[[col for col in section_df.columns if col not in existing_columns]]

                # Ensure only required columns are kept
                final_columns = [col for col in subject_columns + ['Total Marks', 'Rank', 'S.No'] if col in section_df.columns]
                section_df = section_df[final_columns]

                # Remove any completely empty rows
                section_df = section_df.dropna(how='all')

                # Create a new workbook
                wb = Workbook()
                ws = wb.active

                # Add header
                header1 = "VRS & VJ Residential School"
                header2 = f"Exam Marks Report for {section} Periodic Test-1"

                ws.append([header1])
                ws.append([header2])

                # Merge header cells
                ws.merge_cells('A1:L1')
                ws.merge_cells('A2:L2')

                # Apply header formatting
                header_font = Font(bold=True)
                ws['A1'].font = header_font
                ws['A2'].font = header_font

                # Append the DataFrame to the worksheet
                for r_idx, row in enumerate(dataframe_to_rows(section_df, index=False, header=True), 3):
                    ws.append(row)

                # Set alignment for header
                for cell in ws[1]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for cell in ws[2]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Set alignment for data cells
                for row in ws.iter_rows(min_row=3):
                    for cell in row:
                        if cell.column < len(subject_columns) + 8:  # Numeric columns based on column index (including BG Marks)
                            cell.alignment = Alignment(horizontal='center')
                        else:  # Text columns
                            cell.alignment = Alignment(horizontal='left')

                # Set borders for the entire data range
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))

                for row in ws.iter_rows(min_row=3):
                    for cell in row:
                        cell.border = thin_border

                # Add footer: Class Teacher (left), Director (right)
                ws.oddFooter.left.text = "Class Teacher"
                ws.oddFooter.right.text = "Principal"

                # Set page layout to A4, landscape mode, narrow margins
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_margins.left = 0.5
                ws.page_margins.right = 0.5
                ws.page_margins.top = 0.5
                ws.page_margins.bottom = 0.5

                # Autofit column width
                for col_idx, col in enumerate(ws.columns, 1):
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                # Save the workbook
                output_file = os.path.join(file_output_dir, f"{section}.xlsx")
                wb.save(output_file)
                print(f"Saved file: {output_file}")
